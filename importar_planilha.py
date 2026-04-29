from openpyxl import load_workbook
from database import query_one, execute
from config_importacao import MAPEAMENTO_ADMIN


def valor_seguro(v):
    if v is None or v == "":
        return 0
    try:
        return float(v)
    except:
        return 0


def texto_seguro(v):
    if v is None:
        return ""
    return str(v).strip()


def ler_celula(ws, referencia):
    try:
        return ws[referencia].value
    except:
        return None


def importar_planilha(caminho_arquivo, codigo_obra, nome_obra, empresa_id=None):
    """
    Importa dados principais da planilha para o banco:
    - cria/atualiza obra
    - importa custos resumidos por categoria
    - importa medições
    - atualiza campos principais da obra com base em células mapeadas
    """

    wb = load_workbook(caminho_arquivo, data_only=True)

    # --------------------------------------------------
    # 1. GARANTIR A OBRA
    # --------------------------------------------------
    obra = query_one("SELECT * FROM obras WHERE codigo = ?", (codigo_obra,))
    if obra and empresa_id is not None and obra["empresa_id"] != empresa_id:
        raise ValueError("Este codigo de obra ja pertence a outra empresa.")
    if obra:
        obra_id = obra["id"]
        empresa_id_obra = obra["empresa_id"] if "empresa_id" in obra.keys() else empresa_id
    else:
        obra_id = execute(
            """
            INSERT INTO obras (
                empresa_id, codigo, nome, tipologia, status, receita_total,
                orcamento, progresso_percentual
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                empresa_id,
                codigo_obra,
                nome_obra,
                "Casa Térrea",
                "planejamento",
                0,
                0,
                0
            )
        )
        empresa_id_obra = empresa_id

    # --------------------------------------------------
    # 2. REGISTRAR IMPORTAÇÃO
    # --------------------------------------------------
    execute(
        """
        INSERT INTO importacoes (empresa_id, nome_arquivo, observacao)
        VALUES (?, ?, ?)
        """,
        (empresa_id_obra, caminho_arquivo, f"Importação vinculada à obra {codigo_obra}")
    )

    # --------------------------------------------------
    # 3. IMPORTAR CAMPOS GERAIS DA ABA ADMIN
    # --------------------------------------------------
    if "ADMIN" in wb.sheetnames:
        ws = wb["ADMIN"]

        orcamento = valor_seguro(ler_celula(ws, MAPEAMENTO_ADMIN.get("orcamento", "")))
        receita_total = valor_seguro(ler_celula(ws, MAPEAMENTO_ADMIN.get("receita_total", "")))
        progresso_percentual = valor_seguro(ler_celula(ws, MAPEAMENTO_ADMIN.get("progresso_percentual", "")))

        # se o progresso vier em decimal, converte para percentual
        if progresso_percentual <= 1 and progresso_percentual > 0:
            progresso_percentual = progresso_percentual * 100

        execute(
            """
            UPDATE obras
            SET nome = ?, orcamento = ?, receita_total = ?, progresso_percentual = ?
            WHERE id = ?
            """,
            (
                nome_obra,
                orcamento,
                receita_total,
                progresso_percentual,
                obra_id
            )
        )

        categorias_map = {
            "ADMINISTRAÇÃO": "Administração",
            "MÃO DE OBRA": "Mão de Obra",
            "ELETRICISTA": "Eletricista",
            "GESSEIRO": "Gesseiro",
            "PINTOR": "Pintor",
            "MATERIAIS EM GERAL": "Materiais em Geral",
            "ACABAMENTOS": "Acabamentos",
            "FERRAMENTAS": "Ferramentas",
            "EQUIPAMENTOS": "Equipamentos"
        }

        execute("DELETE FROM custos_importados_categoria WHERE obra_id = ?", (obra_id,))

        for linha in range(1, ws.max_row + 1):
            valor_b = texto_seguro(ws[f"B{linha}"].value).upper()

            if valor_b in categorias_map:
                categoria_nome = categorias_map[valor_b]

                subtotal = 0
                for busca in range(linha, min(linha + 120, ws.max_row + 1)):
                    rotulo = texto_seguro(ws[f"B{busca}"].value).upper()
                    valor_c = ws[f"C{busca}"].value

                    if rotulo == "CUSTO TOTAL":
                        subtotal = valor_seguro(valor_c)
                        break

                execute(
                    """
                    INSERT INTO custos_importados_categoria (
                        empresa_id, obra_id, categoria, valor_total, origem
                    )
                    VALUES (?, ?, ?, ?, ?)
                    """,
                    (empresa_id_obra, obra_id, categoria_nome, subtotal, "planilha")
                )

    # --------------------------------------------------
    # 4. IMPORTAR MEDIÇÕES - ABA MEDIÇÕES
    # --------------------------------------------------
    if "MEDIÇÕES" in wb.sheetnames:
        ws = wb["MEDIÇÕES"]

        execute("DELETE FROM medicoes WHERE obra_id = ?", (obra_id,))

        for linha in range(28, 60):
            mes = texto_seguro(ws[f"C{linha}"].value)
            medicao_nome = texto_seguro(ws[f"D{linha}"].value)
            percentual = valor_seguro(ws[f"E{linha}"].value)
            percentual_acumulado = valor_seguro(ws[f"F{linha}"].value)
            valor_realizado = valor_seguro(ws[f"G{linha}"].value)

            if medicao_nome:
                execute(
                    """
                    INSERT INTO medicoes (
                        empresa_id, obra_id, mes, medicao_nome, etapa,
                        percentual, percentual_acumulado,
                        valor_realizado, data_medicao, observacao
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        empresa_id_obra,
                        obra_id,
                        mes,
                        medicao_nome,
                        "Importado da planilha",
                        percentual * 100 if percentual <= 1 and percentual > 0 else percentual,
                        percentual_acumulado * 100 if percentual_acumulado <= 1 and percentual_acumulado > 0 else percentual_acumulado,
                        valor_realizado,
                        "",
                        "Importado automaticamente"
                    )
                )

    return obra_id
